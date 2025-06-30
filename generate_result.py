import pandas as pd
from openpyxl import load_workbook

def normalize(text):
    return str(text).strip().lower().replace('ё', 'е').replace('\u00A0', ' ').replace('_', '').replace('-', '')

def main():
    print("Загрузка данных...")

    stock_file = 'stock_clean.xlsx'
    orders_file = 'orders.csv'
    weight_file = 'Вес.xlsx'

    # Загрузка таблиц
    stock = pd.read_excel(stock_file, sheet_name="Товар-склад", skiprows=2)
    orders = pd.read_csv(orders_file, sep=';', encoding='utf-8', engine='python')
    weight = pd.read_excel(weight_file)

    # Очистка и стандартизация
    stock['Артикул'] = stock['Артикул'].astype(str).str.strip()
    orders['Артикул'] = orders['Артикул'].astype(str).str.strip()
    weight['Артикул'] = weight['Артикул'].astype(str).str.strip()
    orders['Склад отгрузки'] = orders['Склад отгрузки'].astype(str).str.strip()
    stock['Склад'] = stock['Склад'].astype(str).str.strip()

    # Нормализация названий складов
    stock['Склад'] = stock['Склад'].apply(normalize)
    orders['Склад отгрузки'] = orders['Склад отгрузки'].apply(normalize)

    # Считаем продажи по складу и артикулу
    sales = orders.groupby(['Склад отгрузки', 'Артикул'])['Количество'].sum().reset_index()
    sales.rename(columns={'Количество': 'Продано', 'Склад отгрузки': 'Склад'}, inplace=True)

    stock.columns = stock.columns.str.replace('\u200b', '', regex=True)
    orders.columns = orders.columns.str.replace('\u200b', '', regex=True)
    weight.columns = weight.columns.str.replace('\u200b', '', regex=True)

    # Объединение с весом
    merged = sales.merge(weight[['Артикул', 'Вес']], on='Артикул', how='left')

    # Переименуем колонки в stock
    stock.rename(columns={
    'Доступно к продаже': 'Остаток',
    'В заявках на поставку': 'В заявках на поставку',
    'В поставках в пути': 'В пути'
    }, inplace=True)

    # Объединение с остатками и логистикой
    stock_columns = ['Артикул', 'Склад', 'Остаток', 'В заявках на поставку', 'В пути']
    merged = merged.merge(stock[stock_columns], on=['Артикул', 'Склад'], how='left')

    # Заполнение пропусков
    merged[['Продано', 'Остаток', 'В заявках на поставку', 'В пути']] = \
        merged[['Продано', 'Остаток', 'В заявках на поставку', 'В пути']].fillna(0).astype(float)
    merged['Вес'] = merged['Вес'].fillna(0)
    merged = merged.infer_objects(copy=False)

    # Расчёты
    merged['Надо'] = merged['Продано'] * 3
    merged['Отправлять'] = (
        merged['Надо'] - merged['Остаток'] - merged['В заявках на поставку'] - merged['В пути']
    ).clip(lower=0)
    merged['Вес товаров'] = merged['Отправлять'] * merged['Вес']
    merged['Итоговый вес'] = merged['Вес товаров']

    # Добавим вспомогательные колонки из weight справа
    extra_cols = [col for col in weight.columns if col not in ['Артикул', 'Вес', 'Надо', 'Отправлять']]
    extra_data = weight[['Артикул'] + extra_cols]
    merged = merged.merge(extra_data, on='Артикул', how='left')

    # Сохранение двух файлов: выбранный склад и общий отчёт
    result_file_all = 'result_all.xlsx'
    result_file_selected = 'result_selected.xlsx'

    sklads = sorted(merged['Склад'].dropna().unique())

    print("\nДоступные склады:")
    for idx, sklad in enumerate(sklads, 1):
        print(f"{idx}. {sklad}")
    sel_index = int(input("\nВведите номер склада: ")) - 1

    if sel_index < 0 or sel_index >= len(sklads):
        print("Неверный номер склада.")
        return

    selected = sklads[sel_index]
    print(f"Выбран склад: {selected}")

    # Формируем списки колонок
    base_columns = [
    'Артикул', 'Склад',
    'Продано', 'Остаток',
    'В заявках на поставку', 'В пути',
    'Надо', 'Отправлять',
    'Вес', 'Вес товаров', 'Итоговый вес'
    ]
    columns_order = base_columns + extra_cols

    full_df = merged[columns_order]
    selected_df = merged[merged['Склад'] == selected][columns_order]

    # Сохраняем оба файла
    with pd.ExcelWriter(result_file_all, engine='xlsxwriter') as writer:
        full_df.to_excel(writer, index=False, sheet_name='Все склады')

        original_weight = pd.read_excel("Вес.xlsx")
        original_weight.to_excel(writer, index=False, sheet_name="Микс коробка")

    with pd.ExcelWriter(result_file_selected, engine='xlsxwriter') as writer:
        selected_df.to_excel(writer, index=False, sheet_name='Выбранный склад')

    print("\nГотово. Сформированы два файла: result_all.xlsx и result_selected.xlsx")

    wb = load_workbook("result_all.xlsx")
    ws = wb["Микс коробка"]

    # Найти строку с "Коробка"
    box_row = ws.max_row
    for row in range(2, ws.max_row + 1):
        val = str(ws[f"A{row}"].value).strip().lower()
        if val == "коробка":
            box_row = row
            break

    if box_row:
        ws.insert_rows(box_row + 1)
        ws[f"A{box_row + 1}"] = "Итого"
        ws[f"D{box_row + 1}"] = f"=SUM(D2:D{box_row})"
        for row in range(2, box_row):
            ws[f"D{row}"] = f"=B{row}*C{row}"

    wb.save("result_all.xlsx")

if __name__ == "__main__":
    main()
